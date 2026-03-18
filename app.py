"""
HCR Split-Initiator Probe Designer — Streamlit App
====================================================
Upload a FASTA file or paste a sequence, choose your initiator set,
and download a fully formatted Excel workbook with probe QC metrics
and an ordering sheet.
"""

import io
import re
import sys
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.utils import get_column_letter

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="HCR Probe Designer",
    page_icon="🧬",
    layout="wide",
)

# ── Custom CSS ─────────────────────────────────────────────────────────────────
st.markdown("""
<style>
  [data-testid="stAppViewContainer"] { background: #0f1117; }
  [data-testid="stSidebar"]          { background: #161b22; border-right: 1px solid #30363d; }
  h1, h2, h3, label, p, li          { color: #e6edf3 !important; }
  .stTextArea textarea               { background: #161b22 !important; color: #79c0ff !important;
                                       font-family: 'Courier New', monospace !important; font-size: 13px; }
  .metric-card {
      background: linear-gradient(135deg, #1a2332, #1f2d3d);
      border: 1px solid #30363d; border-radius: 10px;
      padding: 18px 22px; text-align: center;
  }
  .metric-card .val { font-size: 2.2rem; font-weight: 700; color: #58a6ff; }
  .metric-card .lbl { font-size: 0.8rem; color: #8b949e; margin-top: 2px; }
  .pass-badge  { background:#1a3a2a; color:#56d364; border:1px solid #2ea043;
                 border-radius:4px; padding:2px 8px; font-size:12px; font-weight:600; }
  .fail-badge  { background:#3d1f1f; color:#f85149; border:1px solid #8b1a1a;
                 border-radius:4px; padding:2px 8px; font-size:12px; }
  div[data-testid="stDataFrame"] { border-radius: 8px; overflow: hidden; }
  .stButton > button {
      background: linear-gradient(135deg, #1a7a8a, #0d5c6e);
      color: white; border: none; border-radius: 6px;
      padding: 10px 24px; font-weight: 600; font-size: 15px;
      transition: all 0.2s;
  }
  .stButton > button:hover { background: linear-gradient(135deg, #20919f, #1a7a8a); transform: translateY(-1px); }
  .stDownloadButton > button {
      background: linear-gradient(135deg, #1e8449, #145a32) !important;
      color: white !important; border: none !important; border-radius: 6px !important;
      padding: 10px 24px !important; font-weight: 600 !important; font-size: 15px !important;
  }
  .section-header {
      background: linear-gradient(90deg, #1a7a8a22, transparent);
      border-left: 3px solid #1a7a8a; padding: 8px 14px;
      border-radius: 0 6px 6px 0; margin: 18px 0 10px 0;
      color: #79c0ff !important; font-weight: 600; font-size: 1.05rem;
  }
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  SCIENCE CORE  (unchanged logic from original script)
# ══════════════════════════════════════════════════════════════════════════════

INITIATORS = {
    "B1": {"P1": "GAGGAGGGCAGCAAACGG",   "P2": "GAAGAGTCTTCCTTTACG",  "S1": "AA", "S2": "TA"},
    "B2": {"P1": "CCTCGTAAATCCTCATCA",   "P2": "ATCATCCAGTAAACCGCC",  "S1": "AA", "S2": "AA"},
    "B3": {"P1": "GTCCCTGCCTCTATATCT",   "P2": "CCACTCAACTTTAACCCG",  "S1": "TT", "S2": "TT"},
    "B4": {"P1": "CCTCAACCTACCTCCAAC",   "P2": "TCTCACCATATTCGCTTC",  "S1": "AA", "S2": "AT"},
    "B5": {"P1": "CTCACTCCCAATCTCTAT",   "P2": "CTACCCTACAAATCCAAT",  "S1": "AA", "S2": "AA"},
}

def parse_fasta(text: str) -> dict:
    records, header, parts = {}, None, []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        if line.startswith(">"):
            if header is not None:
                records[header] = "".join(parts)
            header = line[1:].split()[0]
            parts = []
        else:
            parts.append(line.upper().replace("U", "T"))
    if header is not None:
        records[header] = "".join(parts)
    return records

_COMP = str.maketrans("ATGCatgc", "TACGtacg")

def reverse_complement(seq):  return seq.translate(_COMP)[::-1]
def gc_content(seq):          s = seq.upper(); return (s.count("G")+s.count("C"))/len(s)*100
def tm_basic(seq):            s = seq.upper(); return 2*(s.count("A")+s.count("T"))+4*(s.count("G")+s.count("C"))

def hairpin_score(seq):
    rc, best = reverse_complement(seq), 0
    for i in range(len(seq)-4):
        for j in range(i+5, len(seq)+1):
            if seq[i:j] in rc: best = max(best, j-i)
    return best

def dimer_score(s1, s2):
    rc, best = reverse_complement(s2), 0
    for i in range(len(s1)-4):
        for j in range(i+5, len(s1)+1):
            if s1[i:j] in rc: best = max(best, j-i)
    return best

def secondary_structure_score(target):
    rc = reverse_complement(target)
    return sum(1 for i in range(len(target)-5) if target[i:i+6] in rc)

def passes_filters(gc, hp1, hp2, dim, struct,
                   gc_min=35, gc_max=65, hp_max=6, dim_max=6, struct_max=4):
    flags = []
    if not (gc_min <= gc <= gc_max): flags.append(f"GC {gc:.1f}%")
    if hp1 > hp_max:                  flags.append(f"Hairpin-P1 {hp1}")
    if hp2 > hp_max:                  flags.append(f"Hairpin-P2 {hp2}")
    if dim > dim_max:                  flags.append(f"Dimer {dim}")
    if struct > struct_max:            flags.append(f"Structure {struct}")
    return (len(flags)==0, "; ".join(flags) if flags else "PASS")

def generate_probes(seq, arm_type, step=52, target_len=52):
    ini = INITIATORS[arm_type]
    P1, P2, S1, S2 = ini["P1"], ini["P2"], ini["S1"], ini["S2"]
    rows = []
    for i in range(0, len(seq)-target_len+1, step):
        target = seq[i:i+target_len]
        arm1, arm2 = target[:25], target[-25:]
        probe1 = P1 + S1 + reverse_complement(arm1)
        probe2 = reverse_complement(arm2) + S2 + P2
        gc, hp1, hp2 = gc_content(target), hairpin_score(probe1), hairpin_score(probe2)
        dim, struct = dimer_score(probe1, probe2), secondary_structure_score(target)
        tm1, tm2 = tm_basic(probe1), tm_basic(probe2)
        ok, flag = passes_filters(gc, hp1, hp2, dim, struct)
        if not ok: continue
        rows.append({
            "Probe_#": len(rows)+1, "Status": "PASS", "Fail_reason": "",
            "Target_start": i+1, "Target_end": i+target_len,
            "Target_sequence": target,
            "Arm1_5to3": arm1, "Arm2_5to3": arm2,
            "Probe1_5to3": probe1, "Probe2_5to3": probe2,
            "Probe1_length": len(probe1), "Probe2_length": len(probe2),
            "GC_percent": round(gc,1), "Tm_P1_C": round(tm1,1), "Tm_P2_C": round(tm2,1),
            "Hairpin_P1": hp1, "Hairpin_P2": hp2,
            "Dimer_score": dim, "Structure_score": struct,
            "Initiator_set": arm_type,
        })
    return pd.DataFrame(rows)

def rank_and_select(df, max_probes=30):
    if df.empty: return df
    df = df.copy()
    df["_score"] = (abs(df["GC_percent"]-50)*0.5 + df["Hairpin_P1"]*2 +
                    df["Hairpin_P2"]*2 + df["Dimer_score"]*2 + df["Structure_score"])
    df = df.sort_values("_score").head(max_probes).sort_values("Target_start").reset_index(drop=True)
    df["Probe_#"] = range(1, len(df)+1)
    return df.drop(columns=["_score"])


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT  (unchanged from original)
# ══════════════════════════════════════════════════════════════════════════════

C_TEAL="1A7A8A"; C_TEAL_LITE="D0EDF1"; C_GREEN="1E8449"; C_GREEN_LT="D5F5E3"
C_RED="C0392B";  C_RED_LT="FADBD8";    C_ORANGE="D35400"; C_ORANGE_LT="FDEBD0"
C_GREY="F2F4F4"; C_WHITE="FFFFFF";     C_DARK="1C2833";   C_SPACER="E74C3C"

def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(h): return PatternFill("solid", fgColor=h)
def _center(): return Alignment(horizontal="center", vertical="center")
def _left():   return Alignment(horizontal="left",   vertical="center")

def _rich_probe(probe_seq, initiator_part, spacer, spacer_after):
    nf = InlineFont(rFont="Arial", sz=20, color=C_DARK)
    sf = InlineFont(rFont="Arial", sz=20, color=C_SPACER)
    s_len, i_len = len(spacer), len(initiator_part)
    if not spacer_after:
        return CellRichText(TextBlock(nf, probe_seq[:i_len]),
                            TextBlock(sf, probe_seq[i_len:i_len+s_len]),
                            TextBlock(nf, probe_seq[i_len+s_len:]))
    else:
        cut = len(probe_seq)-i_len-s_len
        return CellRichText(TextBlock(nf, probe_seq[:cut]),
                            TextBlock(sf, probe_seq[cut:cut+s_len]),
                            TextBlock(nf, probe_seq[cut+s_len:]))

def write_design_sheet(ws, df, gene_name, arm_type):
    ws.merge_cells("A1:R1")
    ws["A1"] = f"HCR Probe Design Report  ·  Gene: {gene_name}  ·  Initiator: {arm_type}"
    ws["A1"].font = Font(name="Arial", size=13, bold=True, color=C_WHITE)
    ws["A1"].fill = _fill(C_TEAL)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:D2")
    ws["A2"] = f"Probes passing QC: {len(df)}   |   Oligos to order: {len(df)*2}"
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color=C_DARK)
    ws["A2"].fill = _fill(C_TEAL_LITE)
    ws["A2"].alignment = _left()

    ws.merge_cells("A4:R4")
    ws["A4"] = f"✔  PASSING PROBES  ({len(df)} probes ready for ordering)"
    ws["A4"].font = Font(name="Arial", size=11, bold=True, color=C_WHITE)
    ws["A4"].fill = _fill(C_GREEN)
    ws["A4"].alignment = _left()
    ws.row_dimensions[4].height = 22

    HEADERS = ["#","Start","End","Target Sequence (52 nt)",
               "Arm1 (25 nt)","Arm2 (25 nt)",
               "Probe 1  (5'→3')","Probe 2  (5'→3')",
               "P1 Len","P2 Len","GC %","Tm P1 (°C)","Tm P2 (°C)",
               "Hairpin P1","Hairpin P2","Dimer","Structure","Initiator"]
    COL_WIDTHS = [5,7,7,34,16,16,42,42,7,7,8,10,10,10,10,8,10,10]
    for ci,(h,w) in enumerate(zip(HEADERS,COL_WIDTHS),1):
        cell = ws.cell(row=5, column=ci, value=h)
        cell.font = Font(name="Arial", size=11, bold=True, color=C_WHITE)
        cell.fill = _fill(C_DARK); cell.alignment = _center(); cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[5].height = 20

    ini = INITIATORS[arm_type]
    KEYS = ["Probe_#","Target_start","Target_end","Target_sequence",
            "Arm1_5to3","Arm2_5to3","Probe1_5to3","Probe2_5to3",
            "Probe1_length","Probe2_length","GC_percent","Tm_P1_C","Tm_P2_C",
            "Hairpin_P1","Hairpin_P2","Dimer_score","Structure_score","Initiator_set"]
    for r_off,(_,rec) in enumerate(df.iterrows()):
        row = 6+r_off
        for ci,key in enumerate(KEYS,1):
            val = rec[key]; cell = ws.cell(row=row, column=ci)
            if key == "Probe1_5to3":
                cell.value = _rich_probe(val, ini["P1"], ini["S1"], False)
            elif key == "Probe2_5to3":
                cell.value = _rich_probe(val, ini["P2"], ini["S2"], True)
            else:
                cell.value = val
                cell.font  = Font(name="Arial", size=10, bold=(key=="Probe_#"), color=C_DARK)
            cell.fill = _fill(C_GREEN_LT); cell.border = _thin_border()
            cell.alignment = _left() if ci>3 else _center()
    ws.freeze_panes = ws["A6"]

def write_order_sheet(ws, df, gene_name):
    ws.merge_cells("A1:G1")
    ws["A1"] = f"Oligo Ordering Sheet  ·  {gene_name}  ·  HCR Probes"
    ws["A1"].font = Font(name="Arial", size=13, bold=True, color=C_WHITE)
    ws["A1"].fill = _fill(C_TEAL); ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    ws["A2"] = f"{len(df)} probe pairs  ·  {len(df)*2} oligos total  ·  All 5'→3', unmodified, standard desalting"
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color=C_DARK)
    ws["A2"].fill = _fill(C_TEAL_LITE); ws["A2"].alignment = _left()

    HEADS  = ["Oligo Name","Sequence (5'→3')","Length (nt)","GC %","Tm (°C)","Purification","Notes"]
    WIDTHS = [22,52,12,8,10,14,30]
    for ci,(h,w) in enumerate(zip(HEADS,WIDTHS),1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = Font(name="Arial", size=11, bold=True, color=C_WHITE)
        cell.fill = _fill(C_DARK); cell.alignment = _center(); cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 20

    r = 4
    for idx,rec in df.iterrows():
        ini = INITIATORS[rec["Initiator_set"]]
        for pnum,(sk,lbl) in enumerate([("Probe1_5to3","P1"),("Probe2_5to3","P2")],1):
            seq  = rec[sk]; name = f"{gene_name}_{rec['Initiator_set']}_pair{idx+1:02d}_{lbl}"
            rich = _rich_probe(seq, ini["P1" if lbl=="P1" else "P2"],
                               ini["S1" if lbl=="P1" else "S2"], lbl=="P2")
            vals = [name, rich, len(seq), round(gc_content(seq),1), round(tm_basic(seq),1),
                    "STD desalt", f"Target pos {rec['Target_start']}–{rec['Target_end']} | Arm{pnum}"]
            bg = C_GREEN_LT if pnum==1 else C_GREY
            for ci,v in enumerate(vals,1):
                cell = ws.cell(row=r, column=ci); cell.value = v
                if ci != 2: cell.font = Font(name="Arial", size=10, bold=(ci==1), color=C_DARK)
                cell.fill = _fill(bg); cell.border = _thin_border()
                cell.alignment = _left() if ci in (1,2,7) else _center()
            r += 1
        for ci in range(1,8): ws.cell(row=r,column=ci).fill = _fill(C_WHITE)
        ws.row_dimensions[r].height = 4; r += 1
    ws.freeze_panes = ws["A4"]

def write_info_sheet(ws, gene_name, arm_type, seq_len, step, df):
    ini = INITIATORS[arm_type]
    ws.merge_cells("A1:D1")
    ws["A1"] = "HCR Probe Designer – Run Summary"
    ws["A1"].font = Font(name="Arial", size=13, bold=True, color=C_WHITE)
    ws["A1"].fill = _fill(C_TEAL); ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 28
    info = [
        ("Gene / Sequence name", gene_name), ("Sequence length (nt)", seq_len),
        ("Initiator set", arm_type), ("P1 initiator", ini["P1"]), ("P2 initiator", ini["P2"]),
        ("Spacer S1", ini["S1"]), ("Spacer S2", ini["S2"]),
        ("Tiling step (nt)", step), ("Target arm length (nt)", 25),
        ("Full target span (nt)", 52), ("Max probes cap", 30),
        ("Probes in output", len(df)), ("Oligos to order", len(df)*2),
        ("GC filter","35 % – 65 %"), ("Hairpin filter","≤ 6 nt"),
        ("Dimer filter","≤ 6 nt"), ("Structure filter","≤ 4 matches"),
    ]
    for i,(k,v) in enumerate(info, 2):
        ws.cell(row=i,column=1,value=k).font = Font(name="Arial",size=10,bold=True,color=C_DARK)
        ws.cell(row=i,column=1).fill = _fill(C_TEAL_LITE)
        ws.cell(row=i,column=2,value=v).font = Font(name="Arial",size=10,color=C_DARK)
        ws.cell(row=i,column=2).fill = _fill(C_WHITE)
    ws.column_dimensions["A"].width = 28; ws.column_dimensions["B"].width = 30

def build_excel_bytes(df, gene_name, arm_type, seq_len, step):
    wb = Workbook()
    ws_info   = wb.active; ws_info.title = "Run_Info"
    ws_design = wb.create_sheet("Probe_Design")
    ws_order  = wb.create_sheet("Order_Sheet")
    write_info_sheet(ws_info,   gene_name, arm_type, seq_len, step, df)
    write_design_sheet(ws_design, df, gene_name, arm_type)
    write_order_sheet(ws_order,  df, gene_name)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
#  STREAMLIT UI
# ══════════════════════════════════════════════════════════════════════════════

st.markdown("""
<div style="padding:12px 0 4px 0">
  <span style="font-size:2rem;font-weight:800;color:#58a6ff;letter-spacing:-1px">🧬 HCR Probe Designer</span>
  <span style="font-size:0.9rem;color:#8b949e;margin-left:12px">Split-Initiator · HCR v3.0</span>
</div>
<p style="color:#8b949e;font-size:0.92rem;margin-top:2px">
  Design optimised HCR split-initiator probe pairs from any mRNA sequence.
  Exports a ready-to-order Excel workbook with full QC metrics.
</p>
<hr style="border-color:#30363d;margin:10px 0 18px 0">
""", unsafe_allow_html=True)

# ── Sidebar controls ───────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### ⚙️ Design Parameters")

    arm_type = st.selectbox(
        "Initiator set",
        options=list(INITIATORS.keys()),
        index=0,
        help="Choose the HCR B-series amplifier set. Use different sets for multiplexing."
    )

    ini_info = INITIATORS[arm_type]
    st.markdown(f"""
    <div style="background:#1f2d3d;border:1px solid #30363d;border-radius:8px;padding:10px 14px;margin:8px 0;font-size:12px;color:#8b949e">
      <b style="color:#58a6ff">{arm_type}</b><br>
      P1: <code style="color:#79c0ff">{ini_info['P1']}</code><br>
      P2: <code style="color:#79c0ff">{ini_info['P2']}</code><br>
      Spacer S1: <code style="color:#ffa657">{ini_info['S1']}</code> &nbsp;
      S2: <code style="color:#ffa657">{ini_info['S2']}</code>
    </div>
    """, unsafe_allow_html=True)

    tiling_step = st.slider(
        "Tiling step (nt)", min_value=52, max_value=200, value=52, step=4,
        help="Distance between consecutive probe start positions. 52 = no overlap (recommended)."
    )
    max_probes = st.slider(
        "Max probes", min_value=5, max_value=50, value=30, step=5,
        help="Maximum number of top-scoring probes to return."
    )

    st.markdown("---")
    st.markdown("### 🔬 QC Thresholds")
    gc_min = st.number_input("GC min (%)", value=35, min_value=20, max_value=50)
    gc_max = st.number_input("GC max (%)", value=65, min_value=50, max_value=80)
    hp_max = st.number_input("Hairpin max (nt)", value=6, min_value=3, max_value=12)
    dm_max = st.number_input("Dimer max (nt)",   value=6, min_value=3, max_value=12)
    st_max = st.number_input("Structure max",    value=4, min_value=1, max_value=10)

# ── Sequence input ─────────────────────────────────────────────────────────────
st.markdown('<div class="section-header">📂 Sequence Input</div>', unsafe_allow_html=True)

input_mode = st.radio("Input method", ["Upload FASTA file", "Paste sequence"], horizontal=True)

gene_name, seq = None, None

if input_mode == "Upload FASTA file":
    uploaded = st.file_uploader("Upload a FASTA file", type=["fasta","fa","fna","txt"])
    if uploaded:
        text = uploaded.read().decode("utf-8", errors="ignore")
        records = parse_fasta(text)
        if not records:
            st.error("❌ No sequences found in the uploaded file. Check FASTA format.")
        else:
            gene_name, seq = next(iter(records.items()))
            seq = seq.upper().replace("U","T")
            st.success(f"✅ Loaded **{gene_name}** — {len(seq):,} nt")
else:
    col1, col2 = st.columns([3, 1])
    with col1:
        raw = st.text_area(
            "Paste sequence (FASTA or plain DNA/RNA)",
            height=140,
            placeholder=">your_gene\nATGCATGC...",
        )
    with col2:
        custom_name = st.text_input("Gene name", value="Input_sequence")

    if raw.strip():
        if raw.strip().startswith(">"):
            records = parse_fasta(raw)
            if records:
                gene_name, seq = next(iter(records.items()))
                seq = seq.upper().replace("U","T")
            else:
                st.error("Could not parse FASTA.")
        else:
            seq = re.sub(r"[^ACGTUacgtu]","", raw).upper().replace("U","T")
            gene_name = custom_name or "Input_sequence"

        if seq:
            st.success(f"✅ Sequence ready — {len(seq):,} nt")

# ── Run design ─────────────────────────────────────────────────────────────────
st.markdown("")
run_btn = st.button("🚀  Design Probes", use_container_width=False)

if run_btn:
    if not seq:
        st.error("⚠️  Please provide a sequence first.")
        st.stop()
    if len(seq) < 52:
        st.error(f"⚠️  Sequence too short ({len(seq)} nt). Need at least 52 nt.")
        st.stop()

    with st.spinner("Designing probes and running QC…"):
        df = generate_probes(seq, arm_type, step=tiling_step, target_len=52)

        # Apply custom QC thresholds (re-filter)
        if not df.empty:
            mask = (
                df["GC_percent"].between(gc_min, gc_max) &
                (df["Hairpin_P1"] <= hp_max) &
                (df["Hairpin_P2"] <= hp_max) &
                (df["Dimer_score"] <= dm_max) &
                (df["Structure_score"] <= st_max)
            )
            df = df[mask].reset_index(drop=True)

        if df.empty:
            st.warning("⚠️  No probes passed QC with current settings. "
                       "Try relaxing thresholds or a different initiator set.")
            st.stop()

        df = rank_and_select(df, max_probes=max_probes)

    # ── Metrics ────────────────────────────────────────────────────────────────
    st.markdown('<div class="section-header">📊 Results Summary</div>', unsafe_allow_html=True)

    c1, c2, c3, c4, c5 = st.columns(5)
    metrics = [
        (c1, str(len(df)),          "Probe pairs"),
        (c2, str(len(df)*2),        "Oligos to order"),
        (c3, f"{df['GC_percent'].mean():.1f}%", "Mean GC"),
        (c4, f"{df['Tm_P1_C'].mean():.0f}°C",   "Mean Tm P1"),
        (c5, arm_type,              "Initiator set"),
    ]
    for col, val, lbl in metrics:
        col.markdown(f"""
        <div class="metric-card">
          <div class="val">{val}</div>
          <div class="lbl">{lbl}</div>
        </div>""", unsafe_allow_html=True)

    # ── Probe table ────────────────────────────────────────────────────────────
    st.markdown('<div class="section-header">🧪 Probe QC Table</div>', unsafe_allow_html=True)

    display_cols = ["Probe_#","Target_start","Target_end","GC_percent",
                    "Tm_P1_C","Tm_P2_C","Hairpin_P1","Hairpin_P2",
                    "Dimer_score","Structure_score","Probe1_5to3","Probe2_5to3"]
    display_df = df[display_cols].rename(columns={
        "Probe_#":"#","Target_start":"Start","Target_end":"End",
        "GC_percent":"GC %","Tm_P1_C":"Tm P1","Tm_P2_C":"Tm P2",
        "Hairpin_P1":"HP1","Hairpin_P2":"HP2","Dimer_score":"Dimer",
        "Structure_score":"Struct","Probe1_5to3":"Probe 1","Probe2_5to3":"Probe 2",
    })

    st.dataframe(
        display_df,
        use_container_width=True,
        height=min(400, 50 + len(df)*38),
        hide_index=True,
        column_config={
            "GC %":  st.column_config.ProgressColumn("GC %",  min_value=0, max_value=100, format="%.1f"),
            "Tm P1": st.column_config.NumberColumn("Tm P1 (°C)", format="%.0f°C"),
            "Tm P2": st.column_config.NumberColumn("Tm P2 (°C)", format="%.0f°C"),
        }
    )

    # ── Download ───────────────────────────────────────────────────────────────
    st.markdown('<div class="section-header">⬇️ Export</div>', unsafe_allow_html=True)

    xlsx_bytes = build_excel_bytes(df, gene_name, arm_type, len(seq), tiling_step)

    dl_col, info_col = st.columns([2, 3])
    with dl_col:
        st.download_button(
            label="📥  Download Excel Workbook",
            data=xlsx_bytes,
            file_name=f"HCR_probes_{gene_name}_{arm_type}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with info_col:
        st.info(
            f"Workbook contains **3 sheets**: Run_Info · Probe_Design · Order_Sheet  \n"
            f"Spacers are highlighted in **red** within each probe sequence."
        )

    # ── GC distribution chart ──────────────────────────────────────────────────
    with st.expander("📈 GC Distribution & Tm Chart", expanded=False):
        chart_df = df[["Probe_#","GC_percent","Tm_P1_C","Tm_P2_C"]].set_index("Probe_#")
        col_a, col_b = st.columns(2)
        with col_a:
            st.markdown("**GC % per probe**")
            st.bar_chart(chart_df["GC_percent"], height=220)
        with col_b:
            st.markdown("**Tm per probe (P1 & P2)**")
            st.line_chart(chart_df[["Tm_P1_C","Tm_P2_C"]], height=220)

elif not run_btn:
    st.markdown("""
    <div style="background:#161b22;border:1px dashed #30363d;border-radius:10px;
                padding:32px;text-align:center;color:#8b949e;margin-top:12px">
      <div style="font-size:2.5rem">🧬</div>
      <div style="font-size:1rem;margin-top:8px">
        Enter or upload a sequence, configure parameters in the sidebar, then click
        <b style="color:#58a6ff">Design Probes</b>.
      </div>
    </div>
    """, unsafe_allow_html=True)
